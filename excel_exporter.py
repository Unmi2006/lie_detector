"""
Excel / CSV Exporter for Lie Detector Pro
Exports session data to .xlsx with multiple sheets:
  - Summary
  - Score Timeline
  - Question Results
  - Signal Breakdown
Requires: pip install openpyxl
"""

import os
import time

try:
    import openpyxl
    from openpyxl.styles import (
        PatternFill, Font, Alignment, Border, Side
    )
    from openpyxl.chart import LineChart, BarChart, Reference
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("[ExcelExporter] openpyxl not installed. Run: pip install openpyxl")


# ─── Style helpers ────────────────────────────────────────────────────────────
def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color.lstrip("#"))

def _font(bold=False, color="E2E8F0", size=10):
    return Font(bold=bold, color=color.lstrip("#"), name="Courier New", size=size)

def _center():
    return Alignment(horizontal="center", vertical="center")

def _thin_border():
    s = Side(style="thin", color="1E293B")
    return Border(left=s, right=s, top=s, bottom=s)

HEADER_FILL  = _fill("#06b6d4")
HEADER_FONT  = _font(bold=True, color="0A0E17", size=10)
ROW1_FILL    = _fill("#111827")
ROW2_FILL    = _fill("#0D1520")
ACCENT_FONT  = _font(bold=False, color="06b6d4")

SCORE_FILLS = {
    "TRUTH":     _fill("#22c55e"),
    "UNCERTAIN": _fill("#38bdf8"),
    "SUSPECT":   _fill("#f97316"),
    "DECEPTION": _fill("#ef4444"),
}
SCORE_FONTS = {
    "TRUTH":     _font(bold=True, color="0a0e17"),
    "UNCERTAIN": _font(bold=True, color="0a0e17"),
    "SUSPECT":   _font(bold=True, color="0a0e17"),
    "DECEPTION": _font(bold=True, color="ffffff"),
}

def _label(score):
    if score < 25:  return "TRUTH"
    if score < 50:  return "UNCERTAIN"
    if score < 75:  return "SUSPECT"
    return "DECEPTION"

def _style_header_row(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill   = HEADER_FILL
        cell.font   = HEADER_FONT
        cell.alignment = _center()
        cell.border = _thin_border()

def _style_data_row(ws, row, ncols, even=True):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill   = ROW1_FILL if even else ROW2_FILL
        cell.font   = _font()
        cell.border = _thin_border()


# ─── Main exporter ────────────────────────────────────────────────────────────
def export_excel(report_data: dict, output_dir: str = "reports") -> str:
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl not installed. Run: pip install openpyxl")

    os.makedirs(output_dir, exist_ok=True)
    ts    = time.strftime("%Y%m%d_%H%M%S")
    fname = os.path.join(output_dir, f"lie_detector_{ts}.xlsx")

    wb = openpyxl.Workbook()

    summary   = report_data.get("summary",  {})
    signals   = report_data.get("signals",  {})
    questions = report_data.get("question_results", [])
    history   = report_data.get("score_history",    [])
    sess_time = report_data.get("session_time", "—")
    avg_score = summary.get("avg_score", 0)

    # ── Sheet 1: Summary ──────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Summary"
    ws1.sheet_view.showGridLines = False
    ws1.column_dimensions["A"].width = 28
    ws1.column_dimensions["B"].width = 22

    title_font = Font(bold=True, color="06B6D4", name="Courier New", size=14)
    ws1["A1"] = "LIE DETECTOR PRO — Session Report"
    ws1["A1"].font = title_font
    ws1.merge_cells("A1:B1")
    ws1["A1"].alignment = _center()
    ws1["A1"].fill = _fill("#0A0E17")

    rows = [
        ("Session Date",    sess_time),
        ("Questions Asked", summary.get("total_questions", 0)),
        ("Average Score",   f"{avg_score:.1f} / 100"),
        ("Peak Score",      f"{summary.get('max_score', 0):.1f} / 100"),
        ("Overall Verdict", _label(avg_score)),
        ("", ""),
        ("--- SIGNAL AVERAGES ---", ""),
        ("Visual Score",   f"{signals.get('avg_visual', 0):.1f}"),
        ("Voice Score",    f"{signals.get('avg_voice',  0):.1f}"),
        ("Emotion Score",  f"{signals.get('avg_emotion',0):.1f}"),
        ("Heart Rate Score", f"{signals.get('avg_hr',   0):.1f}"),
    ]
    for i, (k, v) in enumerate(rows, start=3):
        cell_k = ws1.cell(row=i, column=1, value=k)
        cell_v = ws1.cell(row=i, column=2, value=v)
        fill   = ROW1_FILL if i % 2 == 0 else ROW2_FILL
        for cell in (cell_k, cell_v):
            cell.fill   = fill
            cell.font   = _font()
            cell.border = _thin_border()
        cell_k.font = _font(bold=True, color="64748b")

        # Color verdict
        if k == "Overall Verdict":
            lbl = _label(avg_score)
            cell_v.fill = SCORE_FILLS.get(lbl, ROW1_FILL)
            cell_v.font = SCORE_FONTS.get(lbl, _font())
            cell_v.alignment = _center()

    # ── Sheet 2: Score Timeline ───────────────────────────────────────────────
    ws2 = wb.create_sheet("Score Timeline")
    ws2.sheet_view.showGridLines = False
    ws2.column_dimensions["A"].width = 14
    ws2.column_dimensions["B"].width = 14
    ws2.column_dimensions["C"].width = 14

    headers = ["Time (s)", "Score", "Verdict"]
    for c, h in enumerate(headers, 1):
        ws2.cell(row=1, column=c, value=h)
    _style_header_row(ws2, 1, 3)

    t0 = history[0][0] if history else 0
    for i, (t, s) in enumerate(history, start=2):
        lbl = _label(s)
        ws2.cell(row=i, column=1, value=round(t - t0, 1))
        ws2.cell(row=i, column=2, value=s)
        ws2.cell(row=i, column=3, value=lbl)
        _style_data_row(ws2, i, 3, even=(i % 2 == 0))
        # Color verdict
        vc = ws2.cell(row=i, column=3)
        vc.fill = SCORE_FILLS.get(lbl, ROW1_FILL)
        vc.font = SCORE_FONTS.get(lbl, _font())
        vc.alignment = _center()

    # Line chart
    if len(history) > 1:
        chart = LineChart()
        chart.title       = "Score Timeline"
        chart.style       = 10
        chart.y_axis.title = "Score"
        chart.x_axis.title = "Time (s)"
        chart.height, chart.width = 12, 20
        data = Reference(ws2, min_col=2, min_row=1, max_row=len(history)+1)
        chart.add_data(data, titles_from_data=True)
        chart.series[0].graphicalProperties.line.solidFill = "06b6d4"
        chart.series[0].graphicalProperties.line.width     = 20000
        ws2.add_chart(chart, "E2")

    # ── Sheet 3: Question Results ─────────────────────────────────────────────
    ws3 = wb.create_sheet("Questions")
    ws3.sheet_view.showGridLines = False
    ws3.column_dimensions["A"].width = 6
    ws3.column_dimensions["B"].width = 55
    ws3.column_dimensions["C"].width = 12
    ws3.column_dimensions["D"].width = 14

    headers3 = ["#", "Question", "Score", "Verdict"]
    for c, h in enumerate(headers3, 1):
        ws3.cell(row=1, column=c, value=h)
    _style_header_row(ws3, 1, 4)

    for i, r in enumerate(questions, start=2):
        ws3.cell(row=i, column=1, value=i - 1)
        ws3.cell(row=i, column=2, value=r.get("question", ""))
        ws3.cell(row=i, column=3, value=r.get("avg_score", 0))
        ws3.cell(row=i, column=4, value=r.get("label", ""))
        _style_data_row(ws3, i, 4, even=(i % 2 == 0))
        vc = ws3.cell(row=i, column=4)
        vc.fill = SCORE_FILLS.get(r.get("label",""), ROW1_FILL)
        vc.font = SCORE_FONTS.get(r.get("label",""), _font())
        vc.alignment = _center()
        ws3.cell(row=i, column=3).alignment = _center()
        ws3.cell(row=i, column=1).alignment = _center()

    # Bar chart for questions
    if len(questions) > 0:
        bar = BarChart()
        bar.type   = "bar"
        bar.title  = "Score Per Question"
        bar.y_axis.title = "Score"
        bar.height, bar.width = 12, 20
        data = Reference(ws3, min_col=3, min_row=1, max_row=len(questions)+1)
        cats = Reference(ws3, min_col=1, min_row=2, max_row=len(questions)+1)
        bar.add_data(data, titles_from_data=True)
        bar.set_categories(cats)
        bar.series[0].graphicalProperties.solidFill = "06b6d4"
        ws3.add_chart(bar, "F2")

    # ── Sheet 4: Signal Breakdown ─────────────────────────────────────────────
    ws4 = wb.create_sheet("Signals")
    ws4.sheet_view.showGridLines = False
    ws4.column_dimensions["A"].width = 24
    ws4.column_dimensions["B"].width = 14
    ws4.column_dimensions["C"].width = 10
    ws4.column_dimensions["D"].width = 16

    headers4 = ["Signal", "Avg Score", "Weight", "Contribution"]
    for c, h in enumerate(headers4, 1):
        ws4.cell(row=1, column=c, value=h)
    _style_header_row(ws4, 1, 4)

    sig_rows = [
        ("Visual  (eye/blink/head)", signals.get("avg_visual",0),  "35%", signals.get("avg_visual",0)*0.35),
        ("Voice   (pitch/tremor)",   signals.get("avg_voice",0),   "30%", signals.get("avg_voice",0)*0.30),
        ("Emotion (DeepFace)",       signals.get("avg_emotion",0), "20%", signals.get("avg_emotion",0)*0.20),
        ("Heart Rate (rPPG)",        signals.get("avg_hr",0),      "15%", signals.get("avg_hr",0)*0.15),
    ]
    for i, (name, avg, wt, contrib) in enumerate(sig_rows, start=2):
        ws4.cell(row=i, column=1, value=name)
        ws4.cell(row=i, column=2, value=round(avg, 1))
        ws4.cell(row=i, column=3, value=wt)
        ws4.cell(row=i, column=4, value=round(contrib, 1))
        _style_data_row(ws4, i, 4, even=(i % 2 == 0))
        for c in range(2, 5):
            ws4.cell(row=i, column=c).alignment = _center()

    wb.save(fname)
    return fname


def export_csv(report_data: dict, output_dir: str = "reports") -> str:
    """Simple CSV fallback if openpyxl not available."""
    import csv
    os.makedirs(output_dir, exist_ok=True)
    ts    = time.strftime("%Y%m%d_%H%M%S")
    fname = os.path.join(output_dir, f"lie_detector_{ts}.csv")
    history = report_data.get("score_history", [])
    t0 = history[0][0] if history else 0
    with open(fname, "w", newline="") as f:
        w = csv.writer(f)
        w.writerow(["time_s", "score", "verdict"])
        for t, s in history:
            w.writerow([round(t - t0, 1), s, _label(s)])
    return fname
